import pytest
import uuid
from unittest.mock import MagicMock, patch, AsyncMock
from pathlib import Path
from openpyxl import load_workbook

from app.excel import sheet_line_items, sheet_summary, sheet_gst, sheet_flagged
from openpyxl import Workbook


def make_mock_invoice(status="VERIFIED", confidence=0.9, flags=""):
    inv = MagicMock()
    inv.id = uuid.uuid4()
    inv.job_id = uuid.uuid4()
    inv.source_filename = "test_invoice.pdf"
    inv.category = "Raw Material"
    inv.invoice_number = "INV-001"
    inv.invoice_date = "15/01/2024"
    inv.challan_number = None
    inv.document_type = "TAX INVOICE"
    inv.supplier_name = "ABC Castings Pvt Ltd"
    inv.supplier_gstin = "27AABCA1234B1Z5"
    inv.supplier_state = "Maharashtra"
    inv.supplier_address = "123 Industrial Area, Pune"
    inv.supplier_email = "abc@example.com"
    inv.supplier_phone = "9876543210"
    inv.supplier_bank = "HDFC Bank"
    inv.supplier_account_number = "12345678901"
    inv.supplier_ifsc = "HDFC0001234"
    inv.buyer_name = "XYZ Engineering"
    inv.buyer_gstin = "29AABCX5678B1Z3"
    inv.place_of_supply = "Karnataka"
    inv.destination = "Bangalore"
    inv.transport_name = "Bharti Rodways"
    inv.lr_number = "LR12345"
    inv.vehicle_number = "MH12AB1234"
    inv.eway_bill_number = "1234567890"
    inv.irn_number = None
    inv.assessable_value = 10000.0
    inv.tax_type = "IGST"
    inv.igst_percent = 18.0
    inv.igst_amount = 1800.0
    inv.cgst_percent = 0.0
    inv.cgst_amount = 0.0
    inv.sgst_percent = 0.0
    inv.sgst_amount = 0.0
    inv.pf_charges = 0.0
    inv.round_off = 0.0
    inv.grand_total = 11800.0
    inv.total_weight_kg = 50.0
    inv.total_qty = 100
    inv.confidence_score = confidence
    inv.status = status
    inv.flags = flags

    item = MagicMock()
    item.sr_no = 1
    item.die_number = "D001"
    item.po_number = "PO-2024-001"
    item.description = "Cast Iron Bracket"
    item.hsn_sac_code = "7325"
    item.grade = "GG25"
    item.quantity = 100.0
    item.rate = 100.0
    item.amount = 10000.0
    inv.line_items = [item]

    return inv


def test_line_items_sheet_created():
    wb = Workbook()
    wb.remove(wb.active)
    invoices = [make_mock_invoice()]
    sheet_line_items.build_sheet(wb, invoices)
    assert "Line Items" in wb.sheetnames
    ws = wb["Line Items"]
    assert ws.max_row == 2


def test_summary_sheet_created():
    wb = Workbook()
    wb.remove(wb.active)
    invoices = [make_mock_invoice()]
    sheet_summary.build_sheet(wb, invoices)
    assert "Invoice Summary" in wb.sheetnames
    ws = wb["Invoice Summary"]
    assert ws.max_row == 2


def test_gst_sheet_created():
    wb = Workbook()
    wb.remove(wb.active)
    invoices = [make_mock_invoice()]
    sheet_gst.build_sheet(wb, invoices)
    assert "GST Summary" in wb.sheetnames


def test_flagged_sheet_only_shows_flagged():
    wb = Workbook()
    wb.remove(wb.active)
    invoices = [
        make_mock_invoice(status="VERIFIED"),
        make_mock_invoice(status="DUPLICATE", flags="DUPLICATE of abc.pdf"),
        make_mock_invoice(status="NEEDS_REVIEW", confidence=0.5, flags="LOW_CONFIDENCE"),
    ]
    sheet_flagged.build_sheet(wb, invoices)
    ws = wb["Flagged Bills"]
    assert ws.max_row >= 2
