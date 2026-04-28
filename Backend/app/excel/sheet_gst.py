from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List

from app.database import InvoiceORM

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    "Invoice No", "Invoice Date", "Supplier Name", "Supplier GSTIN",
    "Supplier State", "Buyer Name", "Buyer GSTIN", "Place of Supply",
    "Assessable Value", "Tax Type",
    "IGST %", "IGST Amount", "CGST %", "CGST Amount", "SGST %", "SGST Amount",
    "Total Tax", "Grand Total", "ITC Claimable",
    "IRN No", "E-Way Bill No",
]


def build_sheet(wb: Workbook, invoices: List[InvoiceORM]):
    ws = wb.create_sheet("GST Summary")

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    for inv in invoices:
        total_tax = (inv.igst_amount or 0) + (inv.cgst_amount or 0) + (inv.sgst_amount or 0)
        itc_claimable = "YES" if inv.supplier_gstin and inv.buyer_gstin and inv.status == "VERIFIED" else "NO"

        ws.append([
            inv.invoice_number, inv.invoice_date, inv.supplier_name, inv.supplier_gstin,
            inv.supplier_state, inv.buyer_name, inv.buyer_gstin, inv.place_of_supply,
            inv.assessable_value, inv.tax_type,
            inv.igst_percent, inv.igst_amount, inv.cgst_percent, inv.cgst_amount,
            inv.sgst_percent, inv.sgst_amount,
            total_tax, inv.grand_total, itc_claimable,
            inv.irn_number, inv.eway_bill_number,
        ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    widths = [18, 14, 28, 18, 14, 28, 18, 16, 16, 12,
              8, 14, 8, 14, 8, 14, 12, 14, 12, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
