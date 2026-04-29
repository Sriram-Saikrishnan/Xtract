from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List

from app.database import InvoiceORM

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    # Reference
    "Invoice No", "Invoice Date", "Supplier Name",
    # Line Item
    "Sr No", "Die No", "PO No", "Description", "HSN/SAC", "Grade",
    "Qty", "Rate (₹)", "Line Amount (₹)",
    # Invoice-level cost totals
    "Assessable Value", "IGST %", "IGST Amt", "CGST %", "CGST Amt",
    "SGST %", "SGST Amt", "P&F Charges", "Round Off", "Grand Total",
]


def build_sheet(wb: Workbook, invoices: List[InvoiceORM]):
    ws = wb.create_sheet("Items")

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for inv in invoices:
        items = inv.line_items if inv.line_items else [None]

        for item in items:
            ws.append([
                inv.invoice_number, inv.invoice_date, inv.supplier_name,
                item.sr_no if item else None,
                item.die_number if item else None,
                item.po_number if item else None,
                item.description if item else None,
                item.hsn_sac_code if item else None,
                item.grade if item else None,
                item.quantity if item else None,
                item.rate if item else None,
                item.amount if item else None,
                inv.assessable_value, inv.igst_percent, inv.igst_amount,
                inv.cgst_percent, inv.cgst_amount,
                inv.sgst_percent, inv.sgst_amount,
                inv.pf_charges, inv.round_off, inv.grand_total,
            ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    col_widths = [18, 14, 28, 6, 12, 12, 30, 10, 10, 8, 10, 14,
                  14, 8, 12, 8, 12, 8, 12, 10, 10, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
