from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List

from app.database import InvoiceORM

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    # Source
    "Source File", "Category", "Document Type",
    # Invoice Header
    "Invoice No", "Invoice Date", "Challan No",
    # Supplier
    "Supplier Name", "Supplier GSTIN", "Supplier State", "Supplier Address",
    "Supplier Email", "Supplier Phone", "Supplier Bank", "A/C No", "IFSC",
    # Buyer
    "Buyer Name", "Buyer GSTIN", "Place of Supply", "Destination",
    # Logistics
    "Transport", "LR No", "Vehicle No", "E-Way Bill No", "IRN No",
    # Totals
    "Total Items", "Total Qty", "Total Weight (kg)",
    "Assessable Value", "IGST %", "IGST Amt", "CGST %", "CGST Amt",
    "SGST %", "SGST Amt", "P&F Charges", "Round Off", "Grand Total",
    # Quality
    "Confidence", "Status", "Flags",
]


def build_sheet(wb: Workbook, invoices: List[InvoiceORM]):
    ws = wb.create_sheet("Invoice Summary")

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for inv in invoices:
        item_count = len(inv.line_items) if inv.line_items else 0

        ws.append([
            inv.source_filename, inv.category, inv.document_type,
            inv.invoice_number, inv.invoice_date, inv.challan_number,
            inv.supplier_name, inv.supplier_gstin, inv.supplier_state, inv.supplier_address,
            inv.supplier_email, inv.supplier_phone, inv.supplier_bank,
            inv.supplier_account_number, inv.supplier_ifsc,
            inv.buyer_name, inv.buyer_gstin, inv.place_of_supply, inv.destination,
            inv.transport_name, inv.lr_number, inv.vehicle_number,
            inv.eway_bill_number, inv.irn_number,
            item_count, inv.total_qty, inv.total_weight_kg,
            inv.assessable_value, inv.igst_percent, inv.igst_amount,
            inv.cgst_percent, inv.cgst_amount,
            inv.sgst_percent, inv.sgst_amount,
            inv.pf_charges, inv.round_off, inv.grand_total,
            inv.confidence_score, inv.status, inv.flags,
        ])

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    widths = [
        25, 18, 16,
        18, 14, 14,
        28, 18, 14, 30, 20, 14, 20, 16, 12,
        28, 18, 16, 16,
        20, 12, 14, 20, 40,
        10, 10, 14,
        16, 8, 12, 8, 12, 8, 12, 10, 10, 14,
        10, 14, 40,
    ]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
