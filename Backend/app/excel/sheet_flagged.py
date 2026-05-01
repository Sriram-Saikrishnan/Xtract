from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List

from app.database import InvoiceORM
from app.core.gstin_validator import FLAG_REGISTRY, SEVERITY_ORDER

HEADER_FILL = PatternFill("solid", fgColor="C00000")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CRITICAL_FILL = PatternFill("solid", fgColor="FCE4D6")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
INFO_FILL = PatternFill("solid", fgColor="EAF4FB")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    "Source File", "Invoice No", "Invoice Date", "Supplier Name", "Supplier GSTIN",
    "Grand Total", "Confidence Score", "Status", "Severity",
    "Flags", "Flag Details", "Category", "Action Required",
]

FLAGGED_STATUSES = {"NEEDS_REVIEW", "DUPLICATE", "ERROR"}

_ROW_FILLS = {
    "CRITICAL": CRITICAL_FILL,
    "WARNING": WARNING_FILL,
    "INFO": INFO_FILL,
}


def _parse_flags(flags_str: str) -> list[tuple[str, str, str]]:
    """
    Parse the '; '-separated flags string from the DB into
    a list of (code, message, action) tuples.

    Handles:
    - Known flag codes present in FLAG_REGISTRY
    - Legacy DUPLICATE messages: "DUPLICATE of file.jpg (Invoice: INV-001)"
    - Unknown codes: passed through as-is
    """
    if not flags_str:
        return []

    results = []
    for raw in flags_str.split(";"):
        raw = raw.strip()
        if not raw:
            continue

        if raw in FLAG_REGISTRY:
            info = FLAG_REGISTRY[raw]
            results.append((raw, info["message"], info["action"]))
        elif raw.upper().startswith("DUPLICATE"):
            # Dynamic duplicate message stored as full string
            info = FLAG_REGISTRY.get("DUPLICATE", {})
            results.append(("DUPLICATE", raw, info.get("action", "Check for duplicate and remove before filing.")))
        else:
            results.append((raw, raw, "Review manually."))

    return results


def _row_severity(parsed: list[tuple[str, str, str]]) -> str:
    """Return the highest severity among all flags for a row."""
    best = "INFO"
    for code, _msg, _action in parsed:
        flag_severity = FLAG_REGISTRY.get(code, {}).get("severity", "WARNING")
        # Also handle DUPLICATE detected by prefix
        if code == "DUPLICATE":
            flag_severity = "CRITICAL"
        if SEVERITY_ORDER.get(flag_severity, 2) < SEVERITY_ORDER.get(best, 2):
            best = flag_severity
    return best


def build_sheet(wb: Workbook, invoices: List[InvoiceORM]):
    ws = wb.create_sheet("Flagged Bills")

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    flagged = [
        inv for inv in invoices
        if inv.status in FLAGGED_STATUSES or (inv.flags and inv.flags.strip())
    ]

    # Pre-compute severity for each invoice so we can sort
    def _sort_key(inv: InvoiceORM):
        parsed = _parse_flags(inv.flags or "")
        severity = _row_severity(parsed) if parsed else "INFO"
        return (SEVERITY_ORDER.get(severity, 2), inv.confidence_score or 0.0)

    flagged.sort(key=_sort_key)

    for inv in flagged:
        parsed = _parse_flags(inv.flags or "")
        severity = _row_severity(parsed) if parsed else "INFO"

        flag_codes = " | ".join(code for code, _, _ in parsed) if parsed else ""
        flag_details = " | ".join(msg for _, msg, _ in parsed) if parsed else ""
        action_text = " | ".join(action for _, _, action in parsed) if parsed else "Review manually."

        row_data = [
            inv.source_filename,
            inv.invoice_number,
            inv.invoice_date,
            inv.supplier_name,
            inv.supplier_gstin,
            inv.grand_total,
            inv.confidence_score,
            inv.status,
            severity,
            flag_codes,
            flag_details,
            inv.category,
            action_text,
        ]
        ws.append(row_data)

        row_fill = _ROW_FILLS.get(severity, INFO_FILL)
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=ws.max_row, column=col_idx).fill = row_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    widths = [25, 18, 14, 28, 18, 14, 14, 14, 10, 45, 60, 18, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
